"""
reports.py — Report generation for WeighBridge Pro

Generates:
  • Daily / Monthly / By-Vehicle / By-Material summaries
  • CSV, Excel (openpyxl), PDF (fpdf2) exports
  • Matplotlib chart as PNG bytes
"""

import csv
import io
import datetime
import ctypes
from ctypes import wintypes
import os
import tempfile
from pathlib import Path
import db

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
EXCEL_OK = True

from fpdf import FPDF
PDF_OK = True

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
CHART_OK = True


# ─────────────────────────── HELPERS ─────────────────────────────────────────
def _conn():
    return db.get_conn()


def _q(sql, params=()):
    conn = _conn()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ─────────────────────────── REPORT DATA ─────────────────────────────────────
def daily_report(date: str) -> dict:
    """date = 'YYYY-MM-DD'"""
    rows = _q("""
        SELECT * FROM transactions
        WHERE date(entry_time)=? AND status='OUT'
        ORDER BY entry_time
    """, (date,))
    return _summarise(rows, f"Daily Report — {date}")


def monthly_report(year: int, month: int) -> dict:
    prefix = f"{year}-{month:02d}"
    rows = _q("""
        SELECT * FROM transactions
        WHERE entry_time LIKE ? AND status='OUT'
        ORDER BY entry_time
    """, (f"{prefix}%",))
    return _summarise(rows, f"Monthly Report — {prefix}")


def by_vehicle_report(plate: str, from_dt: str, to_dt: str) -> dict:
    rows = _q("""
        SELECT * FROM transactions
        WHERE vehicle_no LIKE ? AND status='OUT'
          AND entry_time BETWEEN ? AND ?
        ORDER BY entry_time
    """, (f"%{plate.upper()}%", from_dt, to_dt))
    return _summarise(rows, f"Vehicle Report — {plate}")


def by_material_report(material: str, from_dt: str, to_dt: str) -> dict:
    rows = _q("""
        SELECT * FROM transactions
        WHERE material LIKE ? AND status='OUT'
          AND entry_time BETWEEN ? AND ?
        ORDER BY entry_time
    """, (f"%{material}%", from_dt, to_dt))
    return _summarise(rows, f"Material Report — {material}")


def _summarise(rows: list, title: str) -> dict:
    total_net   = sum(r.get("net_wt", 0) or 0 for r in rows)
    total_gross = sum(r.get("gross_wt", 0) or 0 for r in rows)
    total_tare  = sum(r.get("tare_wt", 0) or 0 for r in rows)
    total_inv   = sum(r.get("invoice_amt", 0) or 0 for r in rows)
    return {
        "title":       title,
        "rows":        rows,
        "count":       len(rows),
        "total_gross": total_gross,
        "total_tare":  total_tare,
        "total_net":   total_net,
        "total_net_t": db.kg_to_ton(total_net),
        "total_inv":   total_inv,
    }


# ─────────────────────────── CHART ───────────────────────────────────────────
def chart_daily_tonnage(year: int, month: int):
    """Returns PNG bytes of daily tonnage bar chart for a month."""
    if not CHART_OK:
        return None

    prefix = f"{year}-{month:02d}"
    rows = _q("""
        SELECT date(entry_time) AS d, SUM(net_wt)/1000.0 AS tons
        FROM transactions
        WHERE entry_time LIKE ? AND status='OUT'
        GROUP BY d ORDER BY d
    """, (f"{prefix}%",))

    if not rows:
        return None

    days  = [r["d"][-2:] for r in rows]   # day number only
    tons  = [r["tons"] or 0 for r in rows]

    fig, ax = plt.subplots(figsize=(8, 3.5), facecolor="#161b22")
    ax.set_facecolor("#0d1117")
    bars = ax.bar(days, tons, color="#58a6ff", edgecolor="#30363d", linewidth=0.5)

    # Value labels on bars
    for bar, val in zip(bars, tons):
        if val:
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.02 * max(tons, default=1),
                    f"{val:.1f}", ha="center", va="bottom",
                    fontsize=7, color="#8b949e")

    ax.set_title(f"Daily Tonnage — {year}/{month:02d}",
                 color="#e6edf3", fontsize=11, pad=10)
    ax.set_xlabel("Day", color="#8b949e", fontsize=9)
    ax.set_ylabel("Net Tons", color="#8b949e", fontsize=9)
    ax.tick_params(colors="#8b949e", labelsize=8)
    for spine in ax.spines.values():
        spine.set_edgecolor("#30363d")
    plt.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def chart_monthly_tonnage(year: int):
    """Returns PNG bytes of monthly tonnage for a full year."""
    if not CHART_OK:
        return None

    rows = _q("""
        SELECT strftime('%m', entry_time) AS m, SUM(net_wt)/1000.0 AS tons
        FROM transactions
        WHERE entry_time LIKE ? AND status='OUT'
        GROUP BY m ORDER BY m
    """, (f"{year}%",))

    months_map = {r["m"]: r["tons"] or 0 for r in rows}
    labels = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    tons = [months_map.get(f"{i:02d}", 0) for i in range(1, 13)]

    fig, ax = plt.subplots(figsize=(9, 3.5), facecolor="#161b22")
    ax.set_facecolor("#0d1117")
    ax.plot(labels, tons, color="#3fb950", marker="o",
            linewidth=2, markersize=5)
    ax.fill_between(labels, tons, alpha=0.15, color="#3fb950")

    ax.set_title(f"Monthly Tonnage — {year}",
                 color="#e6edf3", fontsize=11, pad=10)
    ax.set_ylabel("Net Tons", color="#8b949e", fontsize=9)
    ax.tick_params(colors="#8b949e", labelsize=8)
    for spine in ax.spines.values():
        spine.set_edgecolor("#30363d")
    plt.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ─────────────────────────── EXPORT CSV ──────────────────────────────────────
HEADERS = ["Ticket No", "Vehicle", "Driver", "Material", "Supplier",
           "Party", "Gross (kg)", "Tare (kg)", "Net (kg)", "Net (ton)", "Munds",
           "Rate", "Invoice", "Weigh-In", "Weigh-Out", "Status", "Operator"]


def _row_to_list(r: dict) -> list:
    net = float(r.get("net_wt", 0) or 0)
    munds_str = f"{int(net // 40.0)}-{int(net % 40.0)}" if net > 0 else "-"
    return [
        r.get("ticket_no",""), r.get("vehicle_no",""), r.get("driver_name",""),
        r.get("material",""), r.get("supplier",""), r.get("party",""),
        r.get("gross_wt",0), r.get("tare_wt",0), r.get("net_wt",0),
        db.kg_to_ton(net), munds_str,
        r.get("rate",0), r.get("invoice_amt",0),
        r.get("weigh_in_time",""), r.get("weigh_out_time",""),
        r.get("status",""), r.get("operator",""),
    ]


def export_csv(records: list, path: str) -> bool:
    try:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(HEADERS)
            for r in records:
                w.writerow(_row_to_list(r))
        return True
    except Exception:
        return False


# ─────────────────────────── EXPORT EXCEL ────────────────────────────────────
def export_excel(records: list, path: str, title: str = "WeighBridge Report") -> bool:
    if not EXCEL_OK:
        return False
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WeighBridge Data"

        # Title row
        ws.merge_cells(f"A1:{chr(64+len(HEADERS))}1")
        tc = ws["A1"]
        tc.value = title
        tc.font  = Font(bold=True, size=13, color="FFFFFF")
        tc.fill  = PatternFill("solid", fgColor="1F6FEB")
        tc.alignment = Alignment(horizontal="center")

        # Header row
        hdr_fill = PatternFill("solid", fgColor="E0E0E0")
        hdr_font = Font(bold=True, color="000000", size=9)
        for ci, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        alt_fill = PatternFill("solid", fgColor="F5F5F5")
        for ri, r in enumerate(records, 3):
            for ci, val in enumerate(_row_to_list(r), 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ri % 2 == 1:
                    cell.fill = alt_fill
                cell.alignment = Alignment(horizontal="center")

        # Auto-width approximation
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        wb.save(path)
        return True
    except Exception as e:
        raise e


# ─────────────────────────── EXPORT PDF ──────────────────────────────────────
def export_pdf(records: list, path: str, title: str = "WeighBridge Report",
               company: str = "WeighBridge Pro") -> bool:
    if not PDF_OK:
        return False
    try:
        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.add_page()

        # Header
        pdf.set_fill_color(31, 111, 235)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, f"{company} - {title}", ln=True, fill=True, align="C")
        pdf.ln(2)

        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 5, f"Generated: {db.now_str()}   |   Records: {len(records)}", ln=True)
        pdf.ln(3)

        # Table header
        col_w = [26, 22, 20, 17, 18, 18, 18, 18, 18, 16, 18, 22, 16]
        short_hdrs = ["Ticket","Vehicle","Driver","Material","Supplier",
                      "Party","Gross kg","Tare kg","Net kg","Net T", "Munds",
                      "Weigh-In","Status"]
        pdf.set_fill_color(230, 230, 230)
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "B", 7)
        for h, w in zip(short_hdrs, col_w):
            pdf.cell(w, 6, h, border=1, fill=True, align="C")
        pdf.ln()

        # Data rows
        pdf.set_font("Helvetica", "", 6.5)
        for i, r in enumerate(records):
            fill = i % 2 == 0
            pdf.set_fill_color(245, 245, 245) if fill else pdf.set_fill_color(255,255,255)
            pdf.set_text_color(20, 20, 20)
            net = float(r.get("net_wt", 0) or 0)
            munds_str = f"{int(net // 40.0)}-{int(net % 40.0)}" if net > 0 else "-"
            vals = [
                str(r.get("ticket_no",""))[:12],
                str(r.get("vehicle_no",""))[:10],
                str(r.get("driver_name",""))[:10],
                str(r.get("material",""))[:10],
                str(r.get("supplier",""))[:10],
                str(r.get("party",""))[:10],
                f"{r.get('gross_wt',0):,.0f}",
                f"{r.get('tare_wt',0):,.0f}",
                f"{net:,.0f}",
                f"{db.kg_to_ton(net):.3f}",
                munds_str,
                str(r.get("weigh_in_time",""))[:16],
                str(r.get("status",""))[:4],
            ]
            for val, w in zip(vals, col_w):
                pdf.cell(w, 5, val, border=1, fill=fill, align="C")
            pdf.ln()

        # Totals
        total_net = sum(r.get("net_wt",0) or 0 for r in records)
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(31, 111, 235)
        pdf.cell(0, 6, f"TOTAL NET WEIGHT: {total_net:,.1f} kg  |  "
                       f"{db.kg_to_ton(total_net):.3f} ton  |  "
                       f"Count: {len(records)}", ln=True)

        pdf.output(path)
        return True
    except Exception as e:
        raise e


# format_ticket_pdf removed


def format_ticket(tx: dict, company: str = "WeighBridge Pro",
                  address: str = "", footer: str = "") -> str:
    """Generates a balanced-width grid-based text ticket for precise alignment."""
    t1 = str(tx.get("weigh_in_time", ""))
    t2 = str(tx.get("weigh_out_time", ""))
    
    # Auto-fill current date/time
    if not t1 and not t2:
        now = datetime.datetime.now()
        display_date = now.strftime("%d-%m-%Y")
        display_time = now.strftime("%I:%M %p")
    else:
        dt_str = t2 if t2 else t1
        try:
            dt_obj = datetime.datetime.strptime(dt_str[:16], "%Y-%m-%d %H:%M")
            display_date = dt_obj.strftime("%d-%m-%Y")
            display_time = dt_obj.strftime("%I:%M %p")
        except:
            display_date = dt_str[:10]
            display_time = dt_str[11:16]
    
    net = tx.get("net_wt", 0) or 0
    tare = tx.get("tare_wt", 0) or 0
    gross = tx.get("gross_wt", 0) or 0
    m40 = f"{int(net // 40.0)}-{int(net % 40.0)}" if net > 0 else "0-0"
    
    p_name = str(tx.get('party',''))[:44]
    m_name = str(tx.get('material',''))[:44]
    v_no   = str(tx.get('vehicle_no',''))[:9]
    t_no   = str(tx.get('ticket_no',''))[:10]

    # SHRUNKEN ALIGNMENT (Iteration 5):
    # Header: 46 chars (14pt Bold)
    # Body:   54 chars (12pt Regular)
    line46 = "+" + "-"*44 + "+"
    line54 = "+" + "-"*52 + "+"
    
    # Body Table Strategy (Sum: 54 chars incl borders)
    # Format: | (1) + 7 + | (1) + 9 + | (1) + 9 + | (1) + 9 + | (1) + 14 + | (1) = 54
    table_div = "+-------+---------+---------+---------+--------------+"
    
    lines = [
        line46,
        f"|{'SAKHAWAT COMPUTERIZED KANTA':^44}|",
        line46,
        f"|{'ADA PULL 59/15.L MIANCHANNU':^44}|",
        line46,
        # Body starts here (EXACTLY 54 chars wide)
        line54,
        f"| Customer Name: | {p_name:<35} |",
        line54,
        f"| Item Name:     | {m_name:<35} |",
        table_div,
        "|VEHICLE| TICKET  |  DATE   |  TIME   |  WEIGHT KG   |",
        table_div,
        f"|{v_no:^7}|{t_no:^9}|{display_date:^9}|{display_time:^9}| 1st WEIGHT  |",
        f"|       |         |         |         | {tare:>10,.0f} |",
        table_div,
        f"|       |         |         |         | 2nd WEIGHT  |",
        f"|       |         |         |         | {gross:>10,.0f} |",
        "+-------------------------------------+--------------+",
        "| Received with Thanks                |  NET WEIGHT  |",
        f"| OPERATOR NAME : QAISER              | {net:>12,.0f} |",
        "+-------------------------------------+--------------+",
        f"|  24 HOURS SERVICE      |MUNDS:@40KG | {m40:>12} |",
        "+---------------+---------------+--------------------+",
        "|   LAL KHAN    |QAISER SHEHZAD |  RANA ABDULLATIF   |",
        "+---------------+---------------+--------------------+",
        "|  03017538658  |  03061020156  |    03005243291     |",
        "+---------------+---------------+--------------------+"
    ]
    return "\n".join(lines)


def format_ticket_rtf(tx: dict) -> str:
    """Generates an RTF ticket with precisely aligned borders across mixed fonts."""
    content = format_ticket(tx)
    lines = content.split("\n")
    
    rtf = r"{\rtf1\ansi\deff0{\fonttbl{\f0\fmodern\fcharset0 Courier New;}}"
    rtf += r"{\colortbl ;\red0\green0\blue0;}"
    rtf += r"\margl144\margr144\margt144\margb144 "
    
    # Default body size 14pt
    rtf += r"\viewkind4\uc1\f0\fs28 " 
    
    for i, line in enumerate(lines):
        line = line.replace("\\", "\\\\").replace("{", "\\{").replace("}", "\\}")
        
        # Lines 0-4 (Header) are 16pt (fs32) and Bold (b)
        if i < 5:
            rtf += rf"\b\fs32 {line}\b0\fs28\par "
        else:
            # Lines 5+ (Body) are 14pt (fs28) and NOT bold
            rtf += rf"{line}\par "
            
    rtf += "}"
    return rtf


def native_print_ticket(printer_name, tx_data, copies=1):
    """Prints the ticket directly using Windows GDI for mixed fonts."""
    content = format_ticket(tx_data)
    lines = content.split("\n")

    # Windows API Constants & Structures
    HORZRES = 8
    VERTRES = 10
    LOGPIXELSY = 90

    class DOCINFO(ctypes.Structure):
        _fields_ = [("cbSize", wintypes.INT),
                    ("lpszDocName", wintypes.LPCWSTR),
                    ("lpszOutput", wintypes.LPCWSTR),
                    ("lpszDatatype", wintypes.LPCWSTR),
                    ("fwType", wintypes.DWORD)]

    class SIZE(ctypes.Structure):
        _fields_ = [("cx", wintypes.LONG), ("cy", wintypes.LONG)]

    gdi32 = ctypes.WinDLL('gdi32', use_last_error=True)
    winspool = ctypes.WinDLL('winspool.drv', use_last_error=True)

    for _ in range(copies):
        hdc = gdi32.CreateDCW("WINSPOOL", printer_name, None, None)
        if not hdc:
            raise Exception(f"Could not create Device Context for printer: {printer_name}")

        try:
            di = DOCINFO()
            di.cbSize = ctypes.sizeof(DOCINFO)
            di.lpszDocName = "WeighBridge Ticket"
            
            if gdi32.StartDocW(hdc, ctypes.byref(di)) <= 0:
                raise Exception("StartDoc failed")
            
            if gdi32.StartPage(hdc) <= 0:
                raise Exception("StartPage failed")

            # DPI & Page Dimensions
            dpi_y = gdi32.GetDeviceCaps(hdc, LOGPIXELSY)
            page_width = gdi32.GetDeviceCaps(hdc, HORZRES)
            
            # Create Fonts
            h_header = -int(14 * dpi_y / 72)
            h_body   = -int(12 * dpi_y / 72)

            font_header = gdi32.CreateFontW(h_header, 0, 0, 0, 700, 0, 0, 0, 0, 0, 0, 0, 0, "Courier New")
            font_body = gdi32.CreateFontW(h_body, 0, 0, 0, 400, 0, 0, 0, 0, 0, 0, 0, 0, "Courier New")

            # Measure body width to center the block
            size = SIZE()
            gdi32.SelectObject(hdc, font_body)
            body_border = "+" + "-"*52 + "+"
            gdi32.GetTextExtentPoint32W(hdc, body_border, len(body_border), ctypes.byref(size))
            
            body_px_width = size.cx
            x_center_margin = max(0, (page_width - body_px_width) // 2)

            y = int(0.2 * dpi_y) # 0.2 inch top margin
            x = x_center_margin
            
            line_spacing_h = int(1.2 * abs(h_header))
            line_spacing_b = int(1.2 * abs(h_body))

            for i, line in enumerate(lines):
                if i < 5:
                    old_font = gdi32.SelectObject(hdc, font_header)
                    # Center the header relative to the body? 
                    # To keep left walls aligned perfectly as before, we use the same 'x'
                    gdi32.TextOutW(hdc, x, y, line, len(line))
                    y += line_spacing_h
                else:
                    old_font = gdi32.SelectObject(hdc, font_body)
                    gdi32.TextOutW(hdc, x, y, line, len(line))
                    y += line_spacing_b
                
                gdi32.SelectObject(hdc, old_font)

            gdi32.DeleteObject(font_header)
            gdi32.DeleteObject(font_body)

            gdi32.EndPage(hdc)
            gdi32.EndDoc(hdc)
        finally:
            gdi32.DeleteDC(hdc)

    return True
